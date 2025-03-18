import asyncio
from openpyxl import load_workbook
from loguru import logger
from src.utils.constants import ACCOUNTS_FILE

# 创建全局异步锁
file_lock = asyncio.Lock()

async def update_account(token: str, field: str, value: str) -> bool:
    """
    安全地更新XLSX文件中账号的字段。

    Args:
        token (str): Discord账号令牌
        field (str): 要更新的字段名称 (DISCORD_TOKEN, PROXY, USERNAME, STATUS)
        value (str): 字段的新值

    Returns:
        bool: 如果更新成功返回True，如果账号未找到返回False
    """
    # 字段名称到列索引的映射
    field_mapping = {
        "DISCORD_TOKEN": 0,    # Discord令牌
        "PROXY": 1,           # 代理
        "USERNAME": 2,        # 用户名
        "STATUS": 3,          # 状态
        "PASSWORD": 4,        # 密码
        "NEW_PASSWORD": 5,    # 新密码
        "NEW_NAME": 6,        # 新名称
        "NEW_USERNAME": 7,    # 新用户名
        "NEW_PROFILE_PICTURE": 8  # 新头像
    }

    if field not in field_mapping:
        logger.error(f"无效的字段名: {field}")
        return False

    column_index = field_mapping[field]

    async with file_lock:  # 使用异步上下文管理器
        try:
            # 使用loop.run_in_executor处理文件阻塞操作
            loop = asyncio.get_event_loop()
            workbook = await loop.run_in_executor(
                None, load_workbook, ACCOUNTS_FILE
            )
            sheet = workbook.active

            # 查找包含指定令牌的行
            token_found = False
            for row_index, row in enumerate(sheet.rows, 1):
                if row[0].value == token:
                    # 更新指定单元格的值
                    sheet.cell(row=row_index, column=column_index + 1, value=value)
                    token_found = True
                    break

            if not token_found:
                logger.error(f"未找到令牌为 {token[:10]}... 的账号")
                await loop.run_in_executor(None, workbook.close)
                return False

            # 保存更改
            await loop.run_in_executor(None, workbook.save, ACCOUNTS_FILE)
            await loop.run_in_executor(None, workbook.close)

            logger.success(f"成功更新账号 {token[:10]}... 的 {field} 字段")
            return True

        except Exception as e:
            logger.error(f"更新账号时出错: {str(e)}")
            return False